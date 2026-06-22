#!/usr/bin/env python3
"""Render audit/features.json -> audit/imjustdex-feature-audit.xlsx

Single canonical spreadsheet tracking every imjustdex.com feature across the
4 phases (catalog -> test -> fix -> re-test). Re-run after editing features.json
to regenerate. Color-coded for glanceability.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "features.json")
OUT = os.path.join(HERE, "imjustdex-feature-audit.xlsx")

# Refined-brutalism-ish palette for status (kept legible, not garish)
FILLS = {
    "PASS":    PatternFill("solid", fgColor="C6E7C9"),  # green
    "FAIL":    PatternFill("solid", fgColor="F4B9B5"),  # red
    "ISSUE":   PatternFill("solid", fgColor="FBE2A8"),  # amber
    "N/A":     PatternFill("solid", fgColor="E2E2DE"),  # grey
    "PENDING": PatternFill("solid", fgColor="FFFFFF"),  # blank
    "":        PatternFill("solid", fgColor="FFFFFF"),
}
SEV_FILLS = {
    "Critical": PatternFill("solid", fgColor="C0211B"),
    "High":     PatternFill("solid", fgColor="E0625B"),
    "Medium":   PatternFill("solid", fgColor="F0C24B"),
    "Low":      PatternFill("solid", fgColor="A9C7E0"),
    "":         PatternFill("solid", fgColor="FFFFFF"),
}
SEV_FONT = {
    "Critical": Font(color="FFFFFF", bold=True, size=10),
    "High":     Font(color="FFFFFF", bold=True, size=10),
    "Medium":   Font(color="1A1A1A", bold=True, size=10),
    "Low":      Font(color="1A1A1A", bold=True, size=10),
    "":         Font(size=10),
}

HEAD_FILL = PatternFill("solid", fgColor="1A1A1A")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="CC0000", bold=True, size=16)
THIN = Side(style="thin", color="C9C9C4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLS = [
    ("id", "ID", 7),
    ("area", "Area", 18),
    ("feature", "Feature", 26),
    ("user_story", "User Story", 52),
    ("expected", "Expected Behaviour (from code)", 64),
    ("source", "Source", 30),
    ("p1", "P1 Catalog", 11),
    ("p2", "P2 Test", 11),
    ("sev", "Severity", 10),
    ("p3", "P3 Fix", 40),
    ("p4", "P4 Re-test", 11),
]

def main():
    with open(DATA) as f:
        doc = json.load(f)
    feats = doc["features"]
    meta = doc.get("meta", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Features"

    # Title band
    ncol = len(COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(1, 1, f"imjustdex.com — Feature Audit  ·  {meta.get('generated','')}")
    t.font = TITLE_FONT
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    # Legend band
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    leg = ws.cell(2, 1, "Status: PASS / FAIL / ISSUE / N/A / PENDING   ·   Severity: Critical > High > Medium > Low   ·   P1 catalog → P2 test → P3 fix → P4 re-test")
    leg.font = Font(italic=True, size=9, color="555555")

    header_row = 3
    for c, (_, label, width) in enumerate(COLS, start=1):
        cell = ws.cell(header_row, c, label)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width

    status_cols = {"p1", "p2", "p4"}
    for r, feat in enumerate(feats, start=header_row + 1):
        for c, (key, _, _) in enumerate(COLS, start=1):
            val = feat.get(key, "")
            cell = ws.cell(r, c, val)
            cell.border = BORDER
            if key in status_cols:
                cell.alignment = CENTER
                cell.fill = FILLS.get(val, FILLS[""])
                if val == "Cataloged":
                    cell.fill = FILLS["PASS"]
                cell.font = Font(size=10, bold=val in ("FAIL", "ISSUE"))
            elif key == "sev":
                cell.alignment = CENTER
                cell.fill = SEV_FILLS.get(val, SEV_FILLS[""])
                cell.font = SEV_FONT.get(val, SEV_FONT[""])
            else:
                cell.alignment = WRAP_TOP
                cell.font = Font(size=10)

    last = header_row + len(feats)
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncol)}{last}"

    # Summary sheet
    s = wb.create_sheet("Summary")
    s.column_dimensions["A"].width = 22
    s.column_dimensions["B"].width = 10
    s.cell(1, 1, "Phase 2 Test Results").font = Font(bold=True, size=13, color="CC0000")
    from collections import Counter
    def norm(v):
        v = (v or "").strip()
        if not v:
            return "PENDING"
        return v.split()[0].rstrip("—").strip() or "PENDING"
    c2 = Counter(norm(f.get("p2")) for f in feats)
    c4 = Counter(norm(f.get("p4")) for f in feats)
    resolved = sum(1 for f in feats if f.get("sev") and norm(f.get("p4")) == "PASS")
    still_open = sum(1 for f in feats if f.get("sev") and norm(f.get("p4")) != "PASS")
    csev = Counter(f.get("sev", "") for f in feats if f.get("sev"))
    row = 3
    for label, counter in (("— P2 (test) —", c2), ("— P4 (re-test) —", c4)):
        s.cell(row, 1, label).font = Font(bold=True)
        row += 1
        for k in ["PASS", "ISSUE", "FAIL", "N/A", "PENDING"]:
            if counter.get(k):
                s.cell(row, 1, k)
                cell = s.cell(row, 2, counter[k])
                cell.fill = FILLS.get(k, FILLS[""])
                row += 1
        row += 1
    s.cell(row, 1, "— Issues by severity (found) —").font = Font(bold=True)
    row += 1
    for k in ["Critical", "High", "Medium", "Low"]:
        if csev.get(k):
            s.cell(row, 1, k)
            cell = s.cell(row, 2, csev[k])
            cell.fill = SEV_FILLS[k]
            cell.font = SEV_FONT[k]
            row += 1
    row += 1
    s.cell(row, 1, "Resolved (P4 PASS)").font = Font(bold=True, color="2E7D32")
    s.cell(row, 2, resolved).fill = FILLS["PASS"]
    row += 1
    s.cell(row, 1, "Still open").font = Font(bold=True, color="C0211B")
    oc = s.cell(row, 2, still_open)
    if still_open:
        oc.fill = FILLS["FAIL"]
    row += 2
    s.cell(row, 1, f"Total features: {len(feats)}").font = Font(bold=True)

    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(feats)} features)")

if __name__ == "__main__":
    main()
